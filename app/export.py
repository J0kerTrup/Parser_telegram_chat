import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
from app.config import ELEX_FILE

logger = logging.getLogger()

def save_to_excel(data, filename=ELEX_FILE):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Статистика чатов"

        # Заголовки
        doc_string = [
            "💬👥 Телеграм Чат",
            "📢📺 Телеграм Канал",
            "🚫 Пропущено",
            "🎤💬 Войс-чат",
            "📡🎥 Трансляция",
            "❌🔒 Не доступно",
            "⚠️ Ошибка",
            "✅ Доступно",
            "❌ Недоступно",
        ]

        doc_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        row_start = 1
        for line in doc_string:
            ws.append([line])
            for cell in ws[row_start]:
                cell.fill = doc_fill
            row_start += 1

        headers = ["Название", "ID группы", "Username", "Количество пользователей",
                   "Можно ли писать сообщения", "Войс", "Тип голосового чата",
                   "Тип чата", "Дата последнего сообщения", "Сообщений за сутки", "Пользователей онлайн"]
        
        ws.append(headers)

        header_fill = PatternFill(start_color="00FFCC00", end_color="00FFCC00", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, cell in enumerate(ws[row_start], start=1): 
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        chat_count = 0
        channel_count = 0
        voice_chat_count = 0
        streaming_count = 0
        can_send_messages_count = 0
        cannot_send_messages_count = 0
        skipped_count = 0   
        total_users = 0  # Всего пользователей (чаты + каналы)
        total_chat_users = 0  # Только чаты
        total_channel_users = 0  # Только каналы
        total_online_users = 0  # Всего онлайн пользователей
        error_count = 0 
        


        for row in data:
            error_count += 1
            chat_type = row[7]  
            chat_voice_type = row[6]  
            can_send_messages = row[4]  
            user_count = int(row[3]) if str(row[3]).isdigit() else 0  # Количество участников
            online_count = int(row[10]) if str(row[10]).isdigit() else 0  # Онлайн пользователи
            total_messages_24h = sum(row[9] for row in data if isinstance(row[9], (int, float)))

            if chat_type == "💬👥":
                chat_count += 1
                total_chat_users += user_count
            elif chat_type == "📢📺":
                channel_count += 1
                total_channel_users += user_count

            if chat_voice_type == "🎤💬":
                voice_chat_count += 1
            elif chat_voice_type == "📡🎥":
                streaming_count += 1

            if can_send_messages == "Да":
                can_send_messages_count += 1
            else:
                cannot_send_messages_count += 1

            if any("🚫" in str(cell) for cell in row):
                skipped_count += 1

            total_users += user_count
            total_online_users += online_count

        ws["C2"] = "💬👥 Телеграм Чат"
        ws["D2"] = chat_count
        ws["C3"] = "📢📺 Телеграм Канал"
        ws["D3"] = channel_count
        ws["C4"] = "🎤💬 Войс-чат"
        ws["D4"] = voice_chat_count
        ws["C5"] = "📡🎥 Трансляция"
        ws["D5"] = streaming_count
        ws["C6"] = "✅ Доступно (можно писать)"
        ws["D6"] = can_send_messages_count
        ws["C7"] = "❌ Недоступно (нельзя писать)"
        ws["D7"] = cannot_send_messages_count
        ws["C8"] = "🚫 Пропущено"
        ws["D8"] = skipped_count
        ws["F2"] = "👥 Общее количество участников"
        ws["G2"] = total_users
        ws["F3"] = "💬 Общее количество участников в чатах"
        ws["G3"] = total_chat_users
        ws["F4"] = "📢 Общее количество участников в каналах"
        ws["G4"] = total_channel_users
        ws["F5"] = "🟢 Онлайн пользователей"
        ws["G5"] = total_online_users
        ws["F6"] = "📩 Сообщений за 24ч"
        ws["G6"] = total_messages_24h
        ws["F7"] = "⚠️ Ошибки"
        ws["G7"] = error_count


        cell_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ["C2", "D2", "C3", "D3", "C4", "D4", "C5", "D5", "C6", "D6", "C7", "D7", "C8", "D8",
                     "F2", "G2", "F3", "G3", "F4", "G4", "F5", "G5", "F6", "G6", "F7", "G7"]:
            ws[cell].fill = cell_fill
            ws[cell].alignment = Alignment(horizontal="center", vertical="center")
            ws[cell].font = Font(bold=True)

        for row in data:
            ws.append(row)

        data_fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")

        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        for row in ws.iter_rows(min_row=row_start + 1, max_row=len(data) + row_start, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.fill = data_fill
                cell.border = border  

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        ws.auto_filter.ref = f"A{row_start}:{openpyxl.utils.get_column_letter(len(headers))}{row_start}"

        wb.save(filename)
        logger.info(f"✅ Данные сохранены в {filename}")

    except Exception as e:
        logger.error(f"❌ Ошибка: {str(e)}")
